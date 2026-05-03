"""
Семейный финансовый Telegram-бот.

Стек:
- python-telegram-bot v20+ (async API). Документация: https://docs.python-telegram-bot.org/
- SQLite (стандартная библиотека Python, sqlite3). Документация: https://docs.python.org/3/library/sqlite3.html

Запуск:
1) Получите токен у @BotFather (официальная инструкция Telegram:
   https://core.telegram.org/bots/features#botfather).
2) Установите зависимости:  pip install -r requirements.txt
3) Экспортируйте токен в переменную окружения TELEGRAM_BOT_TOKEN
   (Linux/macOS:  export TELEGRAM_BOT_TOKEN="123:ABC..."
    Windows PS:   $env:TELEGRAM_BOT_TOKEN="123:ABC...")
4) Запустите:  python bot.py

ВАЖНО: бот хранит данные в локальном файле family_finance.db в той же папке.
Резервируйте этот файл — в нём вся ваша финансовая история.
"""

from __future__ import annotations

import asyncio
import calendar
import io
import json
import logging
import os
import sqlite3
import tempfile
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# Опциональные зависимости (могут быть недоступны, если не установлены).
try:
    from anthropic import Anthropic
except ImportError:  # pragma: no cover
    Anthropic = None  # type: ignore

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None  # type: ignore

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

# ---------------------------------------------------------------------------
# Настройки и логирование
# ---------------------------------------------------------------------------

# Путь к БД: по умолчанию рядом со скриптом (для локального запуска).
# На Railway задайте переменную DB_PATH, например DB_PATH=/data/family_finance.db,
# и подключите Volume на /data — тогда база переживёт редеплои.
DB_PATH = Path(os.environ.get("DB_PATH", Path(__file__).with_name("family_finance.db")))

# ID владельца в Telegram. Только этому пользователю разрешено вызывать /backup,
# и только ему отправляются автоматические бэкапы. Узнать свой ID — команда /myid.
OWNER_TELEGRAM_ID = int(os.environ.get("OWNER_TELEGRAM_ID", "0")) or None

# Период автобэкапа (в днях). По умолчанию — 7. Можно переопределить переменной.
BACKUP_INTERVAL_DAYS = int(os.environ.get("BACKUP_INTERVAL_DAYS", "7"))

# Anthropic Claude API. Если ключ не задан — естественный язык и чтение PDF
# работать не будут, но штатные команды по-прежнему работают.
# Ключ берётся со страницы https://console.anthropic.com/
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
# Модель для парсинга. Haiku 4.5 — самая дешёвая и быстрая.
# Цены: $1/$5 за миллион входных/выходных токенов.
# Источник: https://www.anthropic.com/pricing
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-haiku-4-5-20251001")

# Месячный отчёт: бот сам присылает .xlsx в каждый чат, где есть данные.
# По умолчанию — в последний день месяца в 21:00 по часовому поясу системы.
MONTHLY_REPORT_DAY = int(os.environ.get("MONTHLY_REPORT_DAY", "28"))  # 28 — безопасно для всех месяцев
MONTHLY_REPORT_HOUR = int(os.environ.get("MONTHLY_REPORT_HOUR", "21"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
log = logging.getLogger("family-finance-bot")


# ---------------------------------------------------------------------------
# Слой работы с БД
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS debts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    chat_id         INTEGER NOT NULL,
    creditor        TEXT    NOT NULL,
    principal       REAL    NOT NULL,           -- остаток долга
    annual_rate_pct REAL    NOT NULL DEFAULT 0, -- годовая ставка, %
    monthly_payment REAL    NOT NULL DEFAULT 0,
    term_months     INTEGER,
    notes           TEXT,
    created_at      TEXT    NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS incomes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    chat_id     INTEGER NOT NULL,
    source      TEXT    NOT NULL,
    amount      REAL    NOT NULL,
    is_monthly  INTEGER NOT NULL DEFAULT 1,     -- 1 = регулярный, 0 = разовый
    notes       TEXT,
    created_at  TEXT    NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS expenses (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    chat_id     INTEGER NOT NULL,
    category    TEXT    NOT NULL,
    amount      REAL    NOT NULL,
    is_monthly  INTEGER NOT NULL DEFAULT 1,
    notes       TEXT,
    created_at  TEXT    NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS goals (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    chat_id     INTEGER NOT NULL,
    name        TEXT    NOT NULL,
    target      REAL    NOT NULL,
    saved       REAL    NOT NULL DEFAULT 0,
    deadline    TEXT,                            -- ISO-дата или NULL
    notes       TEXT,
    created_at  TEXT    NOT NULL DEFAULT (datetime('now'))
);
"""


@contextmanager
def db_conn():
    """Контекстный менеджер для соединения с SQLite."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with db_conn() as c:
        c.executescript(SCHEMA)
    log.info("БД инициализирована: %s", DB_PATH)


# ---------------------------------------------------------------------------
# Доменные модели и расчёты
# ---------------------------------------------------------------------------


@dataclass
class Debt:
    id: int
    creditor: str
    principal: float
    annual_rate_pct: float
    monthly_payment: float
    term_months: int | None
    notes: str | None


def fetch_debts(chat_id: int) -> list[Debt]:
    with db_conn() as c:
        rows = c.execute(
            "SELECT * FROM debts WHERE chat_id=? ORDER BY annual_rate_pct DESC",
            (chat_id,),
        ).fetchall()
    return [
        Debt(
            id=r["id"],
            creditor=r["creditor"],
            principal=r["principal"],
            annual_rate_pct=r["annual_rate_pct"],
            monthly_payment=r["monthly_payment"],
            term_months=r["term_months"],
            notes=r["notes"],
        )
        for r in rows
    ]


def total_monthly_income(chat_id: int) -> float:
    with db_conn() as c:
        row = c.execute(
            "SELECT COALESCE(SUM(amount),0) AS s FROM incomes WHERE chat_id=? AND is_monthly=1",
            (chat_id,),
        ).fetchone()
    return float(row["s"])


def total_monthly_expense(chat_id: int) -> float:
    with db_conn() as c:
        row = c.execute(
            "SELECT COALESCE(SUM(amount),0) AS s FROM expenses WHERE chat_id=? AND is_monthly=1",
            (chat_id,),
        ).fetchone()
    return float(row["s"])


def total_monthly_debt_payment(chat_id: int) -> float:
    with db_conn() as c:
        row = c.execute(
            "SELECT COALESCE(SUM(monthly_payment),0) AS s FROM debts WHERE chat_id=?",
            (chat_id,),
        ).fetchone()
    return float(row["s"])


def fetch_goals(chat_id: int) -> list[sqlite3.Row]:
    with db_conn() as c:
        return c.execute(
            "SELECT * FROM goals WHERE chat_id=? ORDER BY created_at",
            (chat_id,),
        ).fetchall()


# ---------------------------------------------------------------------------
# Парсинг аргументов команд
# ---------------------------------------------------------------------------


def parse_kv_args(text: str) -> dict[str, str]:
    """
    Простой парсер вида:
        /add_debt creditor=Sberbank principal=500000 rate=18 payment=15000 term=48 notes="ипотека на квартиру"
    Поддержка значений в кавычках.
    """
    out: dict[str, str] = {}
    tokens: list[str] = []
    buf = ""
    in_q = False
    for ch in text:
        if ch == '"':
            in_q = not in_q
            continue
        if ch == " " and not in_q:
            if buf:
                tokens.append(buf)
                buf = ""
            continue
        buf += ch
    if buf:
        tokens.append(buf)

    for tok in tokens[1:]:  # пропускаем саму команду
        if "=" in tok:
            k, _, v = tok.partition("=")
            out[k.strip().lower()] = v.strip()
    return out


def fnum(x: float) -> str:
    """Аккуратное форматирование числа без хвоста .00 для целых."""
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x)):,}".replace(",", " ")
    return f"{x:,.2f}".replace(",", " ")


def parse_money(s: str | None) -> float:
    """
    Парсит сумму. Принимает и точку, и запятую как разделитель дробной части
    (русскоязычная запись типа '324487,98' тоже работает). Игнорирует пробелы.
    Поднимает ValueError, если строка не похожа на число.
    """
    if s is None or str(s).strip() == "":
        raise ValueError("пустое значение")
    return float(str(s).replace(" ", "").replace(",", "."))


# ---------------------------------------------------------------------------
# Claude API: естественный язык и чтение договоров
# ---------------------------------------------------------------------------

# Системный промпт для парсинга свободного ввода в структурированное действие.
LLM_SYSTEM_PROMPT = """\
Ты — парсер команд семейного финансового бота. Получаешь сообщение пользователя
на русском и возвращаешь СТРОГО JSON одного из видов:

{"action":"add_income","source":"...","amount":<число>,"is_monthly":<0|1>,"notes":"..."}
{"action":"add_expense","category":"...","amount":<число>,"is_monthly":<0|1>,"notes":"..."}
{"action":"add_debt","creditor":"...","principal":<число>,"rate":<число>,"payment":<число>,"term":<целое|null>,"notes":"..."}
{"action":"add_goal","name":"...","target":<число>,"saved":<число>,"deadline":"YYYY-MM-DD|null","notes":"..."}
{"action":"balance"}
{"action":"analyze"}
{"action":"distribute","amount":<число>}
{"action":"none","reason":"..."}

Правила:
- Все числа — десятичные, без пробелов, точкой как разделитель.
- "rate" — годовая процентная ставка в процентах.
- "term" — срок в МЕСЯЦАХ. Если в тексте «4 года» → 48.
- Поля, которые пользователь не указал, опускай (кроме action).
- Если непонятно, что делать — верни {"action":"none","reason":"..."}.
- Никакого текста кроме JSON. Никаких markdown-блоков ```json.
"""

# Системный промпт для извлечения параметров из договора кредита.
LLM_CONTRACT_PROMPT = """\
Ты разбираешь текст российского кредитного договора. Верни СТРОГО JSON:

{"creditor":"<название банка/организации>",
 "principal":<сумма_остатка_или_изначальная_сумма_кредита>,
 "rate":<годовая_ставка_в_процентах>,
 "payment":<ежемесячный_платёж>,
 "term":<срок_в_месяцах_или_null>,
 "notes":"<короткая характеристика, например: ипотека/потребкредит/кредитная карта>",
 "confidence":<0..1>}

Требования:
- Все суммы — числами, без пробелов, точкой как разделитель.
- "confidence" — твоя уверенность (0..1) в извлечении ключевых полей.
- Если поля нет в тексте — поставь null.
- Никакого текста кроме JSON.
"""


def llm_available() -> bool:
    return bool(ANTHROPIC_API_KEY) and Anthropic is not None


_anthropic_client: Any = None


def get_anthropic_client():
    """Ленивая инициализация клиента Claude."""
    global _anthropic_client
    if _anthropic_client is None and llm_available():
        _anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)
    return _anthropic_client


def _strip_json_block(text: str) -> str:
    """Убирает обёртку ```json ... ```, если LLM её всё-таки вернёт."""
    t = text.strip()
    if t.startswith("```"):
        t = t.strip("`")
        # Уберём префикс «json\n» если есть
        if t.lower().startswith("json"):
            t = t[4:]
        t = t.strip()
        # Заключительные тройные кавычки уже убрали через strip("`")
    return t


async def llm_call(system: str, user: str, max_tokens: int = 600) -> str:
    """
    Вызов Claude API в отдельном потоке (SDK синхронный, чтобы не блокировать
    цикл событий бота). Возвращает текст ответа модели.
    """
    if not llm_available():
        raise RuntimeError("ANTHROPIC_API_KEY не задан — LLM недоступен.")
    client = get_anthropic_client()

    def _sync():
        msg = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=max_tokens,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        # SDK возвращает content как список блоков; берём текстовый.
        return "".join(
            block.text for block in msg.content if getattr(block, "type", None) == "text"
        )

    return await asyncio.to_thread(_sync)


async def llm_parse_intent(user_text: str) -> dict:
    """Свободный текст → структурированное действие (dict)."""
    raw = await llm_call(LLM_SYSTEM_PROMPT, user_text, max_tokens=400)
    raw = _strip_json_block(raw)
    return json.loads(raw)


async def llm_parse_contract(contract_text: str) -> dict:
    """Текст договора → словарь с параметрами кредита."""
    # Контракты бывают длинные; обрезаем до ~50K символов, чтобы уложиться в лимиты.
    snippet = contract_text[:50000]
    raw = await llm_call(LLM_CONTRACT_PROMPT, snippet, max_tokens=400)
    raw = _strip_json_block(raw)
    return json.loads(raw)


# ---------------------------------------------------------------------------
# Хэндлеры команд
# ---------------------------------------------------------------------------


HELP_TEXT = (
    "<b>Свободный текст (если включён Claude API):</b>\n"
    "Просто напишите боту своими словами, например:\n"
    "• «Получила зарплату 120 000»\n"
    "• «Кредит в Сбере, остаток 324 487,98, ставка 39,9%, плачу 13 734,41 в месяц 57 месяцев»\n"
    "• «Куда направить 30 000 свободных»\n"
    "Бот распарсит, покажет, что понял, и попросит подтвердить.\n\n"
    "<b>Договоры в PDF:</b> просто отправьте файл — бот извлечёт параметры и попросит подтвердить.\n\n"
    "<b>Команды (работают всегда):</b>\n"
    "/add_debt creditor=&lt;кому&gt; principal=&lt;остаток&gt; rate=&lt;%&gt; payment=&lt;ежемес.&gt; term=&lt;мес.&gt;\n"
    "/add_income source=&lt;откуда&gt; amount=&lt;сумма&gt; monthly=1|0\n"
    "/add_expense category=&lt;категория&gt; amount=&lt;сумма&gt; monthly=1|0\n"
    "/add_goal name=&lt;цель&gt; target=&lt;сумма&gt; saved=&lt;накоплено&gt; deadline=YYYY-MM-DD\n"
    "/list_debts /del_debt &lt;id&gt;\n"
    "/balance — сводка\n"
    "/analyze — полный анализ\n"
    "/distribute &lt;сумма&gt; — план распределения\n"
    "/report — Excel-отчёт прямо сейчас\n\n"
    "<b>Резервные копии:</b>\n"
    "/myid — Telegram ID\n"
    "/backup — копия БД владельцу\n\n"
    "/help — эта справка"
)


async def cmd_start(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    init_db()
    await update.message.reply_text(
        "Привет! Я веду семейный учёт долгов, доходов, расходов и целей.\n\n"
        "Начните с /add_income (зарплаты), /add_expense (регулярные расходы), "
        "/add_debt (долги). Затем /analyze покажет картину, "
        "а /distribute &lt;сумма&gt; подскажет, куда направить лишнее.\n\n"
        + HELP_TEXT,
        parse_mode=ParseMode.HTML,
    )


async def cmd_help(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(HELP_TEXT, parse_mode=ParseMode.HTML)


async def cmd_add_debt(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    args = parse_kv_args(update.message.text)
    try:
        creditor = args["creditor"]
        principal = parse_money(args["principal"])
        rate = parse_money(args["rate"]) if args.get("rate") else 0.0
        payment = parse_money(args["payment"]) if args.get("payment") else 0.0
        term = int(parse_money(args["term"])) if args.get("term") else None
    except KeyError:
        await update.message.reply_text(
            "Нужно как минимум: creditor=... principal=...\n"
            "Пример: /add_debt creditor=Sberbank principal=500000 rate=18 payment=15000 term=48\n"
            "Можно с запятой: principal=324487,98"
        )
        return
    except ValueError as e:
        await update.message.reply_text(
            f"Не понял число: {e}.\n"
            "Используйте точку или запятую, например: principal=324487,98 rate=39,9"
        )
        return
    notes = args.get("notes")

    with db_conn() as c:
        cur = c.execute(
            "INSERT INTO debts(chat_id,creditor,principal,annual_rate_pct,"
            "monthly_payment,term_months,notes) VALUES (?,?,?,?,?,?,?)",
            (update.effective_chat.id, creditor, principal, rate, payment, term, notes),
        )
        new_id = cur.lastrowid
    await update.message.reply_text(f"Долг #{new_id} «{creditor}» добавлен.")


async def cmd_list_debts(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    debts = fetch_debts(update.effective_chat.id)
    if not debts:
        await update.message.reply_text("Долгов нет.")
        return
    lines = ["<b>Ваши долги (по убыванию ставки):</b>"]
    total = 0.0
    for d in debts:
        total += d.principal
        term = f", срок {d.term_months} мес." if d.term_months else ""
        notes = f"\n   <i>{d.notes}</i>" if d.notes else ""
        lines.append(
            f"#{d.id} {d.creditor}: остаток {fnum(d.principal)}, "
            f"ставка {d.annual_rate_pct:g}%, платёж {fnum(d.monthly_payment)}/мес{term}{notes}"
        )
    lines.append(f"\n<b>Итого долгов: {fnum(total)}</b>")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


async def cmd_del_debt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text("Использование: /del_debt <id>")
        return
    try:
        did = int(context.args[0])
    except ValueError:
        await update.message.reply_text("id должен быть числом.")
        return
    with db_conn() as c:
        cur = c.execute(
            "DELETE FROM debts WHERE id=? AND chat_id=?",
            (did, update.effective_chat.id),
        )
    if cur.rowcount:
        await update.message.reply_text(f"Долг #{did} удалён.")
    else:
        await update.message.reply_text("Не нашёл такой долг.")


async def cmd_add_income(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    args = parse_kv_args(update.message.text)
    try:
        source = args["source"]
        amount = parse_money(args["amount"])
    except KeyError:
        await update.message.reply_text(
            "Пример: /add_income source=Зарплата amount=120000 monthly=1\n"
            "Можно с запятой: amount=124500,75"
        )
        return
    except ValueError as e:
        await update.message.reply_text(f"Не понял число: {e}.")
        return
    is_monthly = int(args.get("monthly", "1"))
    with db_conn() as c:
        c.execute(
            "INSERT INTO incomes(chat_id,source,amount,is_monthly,notes) VALUES (?,?,?,?,?)",
            (update.effective_chat.id, source, amount, is_monthly, args.get("notes")),
        )
    await update.message.reply_text(f"Доход «{source}» {fnum(amount)} записан.")
    # Автосовет: если появились свободные деньги — сразу предложим план
    cid = update.effective_chat.id
    free = total_monthly_income(cid) - total_monthly_expense(cid) - total_monthly_debt_payment(cid)
    if free > 0:
        await update.message.reply_text(
            f"💡 Свободно по месячному бюджету: {fnum(free)}.\n"
            "Сейчас покажу, куда это направить лучше всего…"
        )
        context.args = [str(int(free))]
        try:
            await cmd_distribute(update, context)
        except Exception as e:  # noqa: BLE001
            log.exception("auto-advice in /add_income: %s", e)


async def cmd_add_expense(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    args = parse_kv_args(update.message.text)
    try:
        category = args["category"]
        amount = float(args["amount"])
    except KeyError:
        await update.message.reply_text(
            "Пример: /add_expense category=Аренда amount=40000 monthly=1"
        )
        return
    is_monthly = int(args.get("monthly", "1"))
    with db_conn() as c:
        c.execute(
            "INSERT INTO expenses(chat_id,category,amount,is_monthly,notes) VALUES (?,?,?,?,?)",
            (update.effective_chat.id, category, amount, is_monthly, args.get("notes")),
        )
    await update.message.reply_text(f"Расход «{category}» {fnum(amount)} записан.")


async def cmd_add_goal(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    args = parse_kv_args(update.message.text)
    try:
        name = args["name"]
        target = float(args["target"])
    except KeyError:
        await update.message.reply_text(
            "Пример: /add_goal name=\"Подушка безопасности\" target=300000 saved=50000 deadline=2026-12-31"
        )
        return
    saved = float(args.get("saved", 0))
    deadline = args.get("deadline")
    with db_conn() as c:
        c.execute(
            "INSERT INTO goals(chat_id,name,target,saved,deadline,notes) VALUES (?,?,?,?,?,?)",
            (update.effective_chat.id, name, target, saved, deadline, args.get("notes")),
        )
    await update.message.reply_text(f"Цель «{name}» добавлена.")


async def cmd_balance(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_chat.id
    inc = total_monthly_income(uid)
    exp = total_monthly_expense(uid)
    debt_pay = total_monthly_debt_payment(uid)
    free = inc - exp - debt_pay
    text = (
        "<b>Месячный баланс</b>\n"
        f"Доходы:           {fnum(inc)}\n"
        f"Расходы:          {fnum(exp)}\n"
        f"Платежи по долгам: {fnum(debt_pay)}\n"
        f"<b>Свободно:        {fnum(free)}</b>"
    )
    await update.effective_message.reply_text(text, parse_mode=ParseMode.HTML)


# ---------------------------------------------------------------------------
# /analyze — общая картина и рекомендации
# ---------------------------------------------------------------------------


def emergency_fund_target(monthly_expense: float, monthly_debt_payment: float) -> float:
    """
    Целевой размер «подушки безопасности» = 3 месяца обязательных трат.
    Источник методологии (3–6 месяцев): Consumer Financial Protection Bureau,
    «An essential guide to building an emergency fund»,
    https://www.consumerfinance.gov/an-essential-guide-to-building-an-emergency-fund/
    Берём нижнюю границу — 3 месяца — как стартовую цель.
    """
    return 3 * (monthly_expense + monthly_debt_payment)


async def cmd_analyze(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_chat.id
    inc = total_monthly_income(uid)
    exp = total_monthly_expense(uid)
    debt_pay = total_monthly_debt_payment(uid)
    free = inc - exp - debt_pay
    debts = fetch_debts(uid)
    goals = fetch_goals(uid)

    lines = ["<b>Анализ ситуации</b>"]
    lines.append(
        f"Доход {fnum(inc)} − расходы {fnum(exp)} − платежи по долгам "
        f"{fnum(debt_pay)} = свободно {fnum(free)}/мес."
    )

    if inc <= 0:
        lines.append("⚠ Не указаны доходы. Добавьте через /add_income.")
    elif free < 0:
        lines.append(
            "🔴 Дефицит бюджета: расходы и платежи превышают доход. "
            "Без сокращения расходов или роста дохода долги будут расти."
        )
    else:
        share = free / inc * 100 if inc else 0
        lines.append(f"Свободные деньги — {share:.1f}% дохода.")

    if debts:
        total_debt = sum(d.principal for d in debts)
        worst = max(debts, key=lambda d: d.annual_rate_pct)
        lines.append(f"Всего долгов: {fnum(total_debt)} ({len(debts)} шт.).")
        lines.append(
            f"Самая дорогая ставка — у «{worst.creditor}»: {worst.annual_rate_pct:g}%/год. "
            "По методу лавины (avalanche) гасить досрочно надо именно его — это даёт минимум переплаты."
        )
    else:
        lines.append("Долгов нет — отличное положение.")

    ef_target = emergency_fund_target(exp, debt_pay)
    lines.append(f"Резервный фонд: цель ≈ {fnum(ef_target)} (3 мес. обязательных трат).")
    if goals:
        lines.append("\n<b>Цели:</b>")
        for g in goals:
            pct = (g["saved"] / g["target"] * 100) if g["target"] else 0
            dl = f", до {g['deadline']}" if g["deadline"] else ""
            lines.append(
                f"• {g['name']}: {fnum(g['saved'])}/{fnum(g['target'])} ({pct:.0f}%){dl}"
            )

    lines.append(
        "\nИсточник стратегии:\n"
        "• Метод лавины vs. снежного кома — Consumer Financial Protection Bureau, "
        "https://www.consumerfinance.gov/about-us/blog/which-debt-pay-first/\n"
        "• Резерв 3–6 мес. — CFPB, "
        "https://www.consumerfinance.gov/an-essential-guide-to-building-an-emergency-fund/"
    )
    await update.effective_message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


# ---------------------------------------------------------------------------
# /distribute <сумма>
# ---------------------------------------------------------------------------


def avalanche_savings(debt: Debt, extra: float) -> float:
    """
    Оценка «грубой выгоды» досрочного погашения за 12 месяцев:
    extra * rate. Это упрощённая верхняя оценка — реальная экономия зависит
    от схемы погашения (аннуитет/дифференцированный) и оставшегося срока.
    Помечено как ориентировочный показатель.
    """
    return extra * debt.annual_rate_pct / 100.0


async def cmd_distribute(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_chat.id
    if not context.args:
        await update.effective_message.reply_text("Использование: /distribute <сумма>")
        return
    try:
        amount = float(context.args[0].replace(",", "."))
    except ValueError:
        await update.effective_message.reply_text("Сумма должна быть числом.")
        return
    if amount <= 0:
        await update.effective_message.reply_text("Сумма должна быть положительной.")
        return

    inc = total_monthly_income(uid)
    exp = total_monthly_expense(uid)
    debt_pay = total_monthly_debt_payment(uid)
    monthly_oblig = exp + debt_pay
    ef_target = emergency_fund_target(exp, debt_pay)

    # Сколько «накоплено» в резерв оцениваем по сумме saved у цели с именем,
    # содержащим «подуш» / «резерв» / «emergency». Если такой цели нет — считаем 0.
    goals = fetch_goals(uid)
    ef_saved = 0.0
    ef_goal_name = None
    for g in goals:
        nm = (g["name"] or "").lower()
        if any(k in nm for k in ("подуш", "резерв", "emergency", "safety")):
            ef_saved = float(g["saved"])
            ef_goal_name = g["name"]
            break

    debts = fetch_debts(uid)
    lines = [f"<b>Куда направить {fnum(amount)}?</b>"]

    # Шаг 1: дефицит бюджета
    if inc and inc - monthly_oblig < 0:
        deficit = monthly_oblig - inc
        lines.append(
            f"⚠ В бюджете дефицит {fnum(deficit)}/мес. Сначала закройте дефицит этого месяца."
        )

    # Шаг 2: резервный фонд
    plan: list[tuple[str, float, str]] = []  # (название, сумма, обоснование)
    remaining = amount

    if ef_saved < ef_target and remaining > 0:
        need_ef = min(ef_target - ef_saved, remaining)
        plan.append(
            (
                f"Резервный фонд{f' ({ef_goal_name})' if ef_goal_name else ''}",
                need_ef,
                f"до цели {fnum(ef_target)} осталось {fnum(ef_target - ef_saved)}",
            )
        )
        remaining -= need_ef

    # Шаг 3: дорогие долги (avalanche)
    if remaining > 0 and debts:
        # Гасим по убыванию ставки, но не больше остатка долга
        for d in debts:
            if remaining <= 0:
                break
            if d.annual_rate_pct <= 0:
                continue  # беспроцентные — позже
            pay = min(d.principal, remaining)
            est = avalanche_savings(d, pay)
            plan.append(
                (
                    f"Досрочно: «{d.creditor}» (#{d.id}, {d.annual_rate_pct:g}%)",
                    pay,
                    f"ориент. экономия за год ≈ {fnum(est)}",
                )
            )
            remaining -= pay

    # Шаг 4: беспроцентные долги
    if remaining > 0 and debts:
        for d in debts:
            if remaining <= 0:
                break
            if d.annual_rate_pct > 0:
                continue
            pay = min(d.principal, remaining)
            plan.append(
                (
                    f"Беспроцентный долг: «{d.creditor}» (#{d.id})",
                    pay,
                    "освобождает ежемесячный поток",
                )
            )
            remaining -= pay

    # Шаг 5: цели (кроме подушки)
    if remaining > 0:
        other_goals = [
            g for g in goals
            if not any(k in (g["name"] or "").lower() for k in ("подуш", "резерв", "emergency", "safety"))
        ]
        for g in other_goals:
            if remaining <= 0:
                break
            need = max(0.0, float(g["target"]) - float(g["saved"]))
            if need <= 0:
                continue
            pay = min(need, remaining)
            plan.append(
                (
                    f"Цель: «{g['name']}»",
                    pay,
                    f"до цели остаётся {fnum(need - pay)}",
                )
            )
            remaining -= pay

    if remaining > 0:
        plan.append(
            (
                "Свободный остаток / инвестиции",
                remaining,
                "все приоритеты закрыты — можно рассмотреть инвестиционный счёт",
            )
        )

    if not plan:
        lines.append("Не нашёл, куда направить — добавьте долги/цели через /add_debt и /add_goal.")
    else:
        for i, (title, sum_, note) in enumerate(plan, 1):
            lines.append(f"{i}. <b>{fnum(sum_)}</b> → {title}\n   <i>{note}</i>")

    lines.append(
        "\n<b>Логика приоритетов:</b>\n"
        "1) Резерв 3 мес. трат — CFPB.\n"
        "2) Долги по убыванию ставки (метод лавины) — CFPB.\n"
        "3) Беспроцентные долги — освобождают денежный поток.\n"
        "4) Финансовые цели.\n"
        "5) Остаток — инвестиции/сбережения.\n"
        "\n<i>Оценка экономии — ориентировочная (extra × ставка). "
        "Точная выгода зависит от типа платежа (аннуитет/дифф.) и оставшегося срока кредита.</i>"
    )
    await update.effective_message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


# ---------------------------------------------------------------------------
# Резервные копии (Уровень 2)
# ---------------------------------------------------------------------------


def make_backup_copy() -> Path:
    """
    Делает безопасную копию SQLite-базы через официальный online-backup API.
    Метод Connection.backup() корректно копирует БД, даже если в неё параллельно
    идут записи. Документация: https://docs.python.org/3/library/sqlite3.html#sqlite3.Connection.backup
    Возвращает путь к временному файлу-копии (вызывающая сторона сама удаляет).
    """
    tmp_dir = Path(tempfile.gettempdir())
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dst_path = tmp_dir / f"family_finance_backup_{stamp}.db"
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(dst_path)
    try:
        with dst:
            src.backup(dst)
    finally:
        src.close()
        dst.close()
    return dst_path


async def cmd_myid(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    """Показывает Telegram user_id и chat_id текущего чата.

    user_id  — нужен для OWNER_TELEGRAM_ID (право на /backup и автобэкапы).
    chat_id  — идентификатор бюджета в данном чате (личного или группового).
    """
    uid = update.effective_user.id
    cid = update.effective_chat.id
    name = update.effective_user.full_name
    chat_type = update.effective_chat.type  # 'private' | 'group' | 'supergroup' | 'channel'
    await update.message.reply_text(
        f"Ваш Telegram <b>user_id</b>: <code>{uid}</code>\n"
        f"Имя: {name}\n"
        f"<b>chat_id</b> текущего чата: <code>{cid}</code> ({chat_type})\n\n"
        "<i>user_id</i> используется для <code>OWNER_TELEGRAM_ID</code> (бэкапы).\n"
        "<i>chat_id</i> — это «семейный» бюджет, который видят все участники чата.\n"
        "В личке с ботом chat_id совпадает с user_id.",
        parse_mode=ParseMode.HTML,
    )


async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Отправляет файл базы владельцу. Доступно только OWNER_TELEGRAM_ID."""
    if OWNER_TELEGRAM_ID is None:
        await update.message.reply_text(
            "Бэкап не настроен: не задан OWNER_TELEGRAM_ID. "
            "Узнайте свой ID через /myid и пропишите его в переменные окружения."
        )
        return
    if update.effective_user.id != OWNER_TELEGRAM_ID:
        await update.message.reply_text("Эта команда доступна только владельцу бота.")
        return
    if not DB_PATH.exists():
        await update.message.reply_text("База ещё пуста — резервировать нечего.")
        return

    tmp_path = make_backup_copy()
    try:
        with open(tmp_path, "rb") as f:
            await context.bot.send_document(
                chat_id=OWNER_TELEGRAM_ID,
                document=f,
                filename=tmp_path.name,
                caption="Резервная копия семейной финансовой БД.",
            )
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


async def auto_backup_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Периодический автобэкап, запускается JobQueue раз в BACKUP_INTERVAL_DAYS дней."""
    if OWNER_TELEGRAM_ID is None or not DB_PATH.exists():
        return
    try:
        tmp_path = make_backup_copy()
    except Exception as e:  # noqa: BLE001
        log.exception("Автобэкап: не удалось скопировать БД: %s", e)
        return
    try:
        with open(tmp_path, "rb") as f:
            await context.bot.send_document(
                chat_id=OWNER_TELEGRAM_ID,
                document=f,
                filename=tmp_path.name,
                caption=f"Автобэкап (каждые {BACKUP_INTERVAL_DAYS} дн.).",
            )
        log.info("Автобэкап отправлен владельцу %s", OWNER_TELEGRAM_ID)
    except Exception as e:  # noqa: BLE001
        log.exception("Автобэкап: не удалось отправить файл: %s", e)
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Естественный язык: парсинг + подтверждение через inline-кнопки
# ---------------------------------------------------------------------------


def _format_pending_action(action: dict) -> str:
    """Человекочитаемое превью действия для подтверждения."""
    a = action.get("action")
    if a == "add_income":
        return (
            "💰 <b>Добавить доход</b>\n"
            f"Источник: {action.get('source','—')}\n"
            f"Сумма: {fnum(float(action.get('amount', 0)))}\n"
            f"Регулярный: {'да' if action.get('is_monthly', 1) else 'разовый'}"
            + (f"\nЗаметка: {action['notes']}" if action.get("notes") else "")
        )
    if a == "add_expense":
        return (
            "💸 <b>Добавить расход</b>\n"
            f"Категория: {action.get('category','—')}\n"
            f"Сумма: {fnum(float(action.get('amount', 0)))}\n"
            f"Регулярный: {'да' if action.get('is_monthly', 1) else 'разовый'}"
            + (f"\nЗаметка: {action['notes']}" if action.get("notes") else "")
        )
    if a == "add_debt":
        term = action.get("term")
        lines = [
            "🏦 <b>Добавить долг</b>",
            f"Кредитор: {action.get('creditor','—')}",
            f"Остаток: {fnum(float(action.get('principal', 0)))}",
            f"Ставка: {action.get('rate', 0)}% годовых",
            f"Платёж: {fnum(float(action.get('payment', 0)))}/мес",
        ]
        if term:
            lines.append(f"Срок: {term} мес.")
        if action.get("notes"):
            lines.append(f"Заметка: {action['notes']}")
        return "\n".join(lines)
    if a == "add_goal":
        return (
            "🎯 <b>Добавить цель</b>\n"
            f"Название: {action.get('name','—')}\n"
            f"Цель: {fnum(float(action.get('target', 0)))}\n"
            f"Уже накоплено: {fnum(float(action.get('saved', 0)))}"
            + (f"\nДедлайн: {action['deadline']}" if action.get("deadline") else "")
        )
    if a == "balance":
        return "📊 Показать сводку месяца"
    if a == "analyze":
        return "🔍 Полный анализ ситуации"
    if a == "distribute":
        return f"💡 Распределить {fnum(float(action.get('amount', 0)))} по приоритетам"
    return f"Действие: {a}"


def _save_pending(context: ContextTypes.DEFAULT_TYPE, action: dict) -> str:
    """Кладёт действие в chat_data под уникальным id и возвращает id."""
    pending = context.chat_data.setdefault("_pending", {})
    aid = uuid.uuid4().hex[:8]
    pending[aid] = action
    return aid


def _pop_pending(context: ContextTypes.DEFAULT_TYPE, aid: str) -> dict | None:
    pending = context.chat_data.get("_pending", {})
    return pending.pop(aid, None)


async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Обработчик свободного текста. Если включён LLM — парсит и просит подтверждения.
    Если LLM выключен — просит использовать команды.
    """
    text = update.message.text or ""
    if text.startswith("/"):
        return  # команды обрабатывает CommandHandler

    if not llm_available():
        await update.message.reply_text(
            "Не понимаю свободный текст без подключённого Claude API.\n"
            "Используйте команды (/help) или попросите администратора задать "
            "переменную <code>ANTHROPIC_API_KEY</code> на Railway.",
            parse_mode=ParseMode.HTML,
        )
        return

    # Сообщение «думает»
    await update.message.chat.send_action("typing")
    try:
        action = await llm_parse_intent(text)
    except json.JSONDecodeError as e:
        await update.message.reply_text(
            f"Не смог разобрать ответ модели как JSON: {e}.\n"
            "Попробуйте сформулировать иначе или используйте команды (/help)."
        )
        return
    except Exception as e:  # noqa: BLE001
        log.exception("LLM intent error: %s", e)
        await update.message.reply_text(
            "Ошибка обращения к Claude API. Подробности в логах.\n"
            "Можно использовать команды (/help) — они работают без LLM."
        )
        return

    if action.get("action") == "none":
        await update.message.reply_text(
            "Не понял, что нужно сделать. " + (action.get("reason") or "")
            + "\nПопробуйте конкретнее или /help."
        )
        return

    aid = _save_pending(context, action)
    preview = _format_pending_action(action)
    keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Сохранить", callback_data=f"do:{aid}")],
            [InlineKeyboardButton("❌ Отмена", callback_data=f"cancel:{aid}")],
        ]
    )
    await update.message.reply_text(
        "Я понял так:\n\n" + preview + "\n\nСохранить?",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Обработка нажатий на inline-кнопки подтверждения."""
    q = update.callback_query
    await q.answer()
    data = q.data or ""
    op, _, aid = data.partition(":")
    action = _pop_pending(context, aid)
    if not action:
        await q.edit_message_text("Это действие уже обработано или устарело.")
        return

    if op == "cancel":
        await q.edit_message_text("Отменено.")
        return

    if op != "do":
        return

    cid = update.effective_chat.id
    a = action.get("action")
    try:
        if a == "add_income":
            with db_conn() as c:
                c.execute(
                    "INSERT INTO incomes(chat_id,source,amount,is_monthly,notes) VALUES (?,?,?,?,?)",
                    (
                        cid,
                        action.get("source", "—"),
                        float(action.get("amount", 0)),
                        int(action.get("is_monthly", 1)),
                        action.get("notes"),
                    ),
                )
            await q.edit_message_text("✅ Доход сохранён.")
            await _maybe_offer_advice(update, context)
            return
        if a == "add_expense":
            with db_conn() as c:
                c.execute(
                    "INSERT INTO expenses(chat_id,category,amount,is_monthly,notes) VALUES (?,?,?,?,?)",
                    (
                        cid,
                        action.get("category", "—"),
                        float(action.get("amount", 0)),
                        int(action.get("is_monthly", 1)),
                        action.get("notes"),
                    ),
                )
            await q.edit_message_text("✅ Расход сохранён.")
            return
        if a == "add_debt":
            term = action.get("term")
            with db_conn() as c:
                c.execute(
                    "INSERT INTO debts(chat_id,creditor,principal,annual_rate_pct,"
                    "monthly_payment,term_months,notes) VALUES (?,?,?,?,?,?,?)",
                    (
                        cid,
                        action.get("creditor", "—"),
                        float(action.get("principal", 0)),
                        float(action.get("rate", 0)),
                        float(action.get("payment", 0)),
                        int(term) if term else None,
                        action.get("notes"),
                    ),
                )
            await q.edit_message_text("✅ Долг сохранён.")
            return
        if a == "add_goal":
            with db_conn() as c:
                c.execute(
                    "INSERT INTO goals(chat_id,name,target,saved,deadline,notes) VALUES (?,?,?,?,?,?)",
                    (
                        cid,
                        action.get("name", "—"),
                        float(action.get("target", 0)),
                        float(action.get("saved", 0)),
                        action.get("deadline"),
                        action.get("notes"),
                    ),
                )
            await q.edit_message_text("✅ Цель сохранена.")
            return
        if a == "balance":
            await q.edit_message_text("Запускаю /balance…")
            await cmd_balance(update, context)
            return
        if a == "analyze":
            await q.edit_message_text("Запускаю /analyze…")
            await cmd_analyze(update, context)
            return
        if a == "distribute":
            context.args = [str(action.get("amount", 0))]
            await q.edit_message_text("Считаю распределение…")
            await cmd_distribute(update, context)
            return
        await q.edit_message_text(f"Неизвестное действие: {a}")
    except Exception as e:  # noqa: BLE001
        log.exception("callback execute error: %s", e)
        await q.edit_message_text(f"Ошибка при сохранении: {e}")


async def _maybe_offer_advice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Автосовет после сохранения дохода через свободный текст. Если свободные
    деньги по месячному бюджету > 0 — бот сам показывает план распределения.
    """
    cid = update.effective_chat.id
    free = total_monthly_income(cid) - total_monthly_expense(cid) - total_monthly_debt_payment(cid)
    if free <= 0:
        return
    await context.bot.send_message(
        chat_id=cid,
        text=(
            f"💡 По итогам месяца у вас свободно {fnum(free)}.\n"
            "Сейчас покажу, куда это направить лучше всего…"
        ),
    )
    context.args = [str(int(free))]
    try:
        await cmd_distribute(update, context)
    except Exception as e:  # noqa: BLE001
        log.exception("auto-advice error: %s", e)


# ---------------------------------------------------------------------------
# PDF-документы: чтение договоров кредита
# ---------------------------------------------------------------------------


def _extract_pdf_text(file_path: Path) -> str:
    """Извлекает текст из PDF. Для сканов (без текстового слоя) вернёт мало или ничего."""
    if PdfReader is None:
        raise RuntimeError("pypdf не установлен.")
    reader = PdfReader(str(file_path))
    parts: list[str] = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception as e:  # noqa: BLE001
            log.warning("PDF: ошибка извлечения страницы: %s", e)
    return "\n".join(parts).strip()


async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Обработчик документов (в основном PDF-договоров).
    Скачивает файл, извлекает текст, отдаёт LLM на парсинг параметров кредита,
    показывает превью и просит подтверждение.
    """
    doc = update.message.document
    if doc is None:
        return
    fname = (doc.file_name or "").lower()
    is_pdf = fname.endswith(".pdf") or doc.mime_type == "application/pdf"
    if not is_pdf:
        await update.message.reply_text(
            "Я умею читать только PDF-документы. Пришлите договор в PDF."
        )
        return
    if not llm_available():
        await update.message.reply_text(
            "Чтение договоров требует Claude API. Установите ANTHROPIC_API_KEY."
        )
        return

    await update.message.chat.send_action("typing")
    # Скачиваем во временный файл
    tmp_dir = Path(tempfile.gettempdir())
    tmp_path = tmp_dir / f"contract_{uuid.uuid4().hex}.pdf"
    try:
        tg_file = await doc.get_file()
        await tg_file.download_to_drive(custom_path=tmp_path)
        text = _extract_pdf_text(tmp_path)
    except Exception as e:  # noqa: BLE001
        log.exception("PDF download/extract error: %s", e)
        await update.message.reply_text(f"Не удалось прочитать PDF: {e}")
        return
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass

    if len(text) < 200:
        await update.message.reply_text(
            "Из PDF извлеклось слишком мало текста — возможно, это скан без текстового слоя.\n"
            "Я могу читать только PDF с текстом (не картинки). "
            "Если у вас скан — можно прогнать его через распознавание (например, "
            "https://www.adobe.com/acrobat/online/ocr-pdf.html ) и прислать снова."
        )
        return

    try:
        info = await llm_parse_contract(text)
    except json.JSONDecodeError as e:
        await update.message.reply_text(
            f"Модель не вернула корректный JSON: {e}. Попробуйте позже."
        )
        return
    except Exception as e:  # noqa: BLE001
        log.exception("LLM contract error: %s", e)
        await update.message.reply_text(f"Ошибка обращения к Claude: {e}")
        return

    confidence = info.get("confidence", 0)
    action = {
        "action": "add_debt",
        "creditor": info.get("creditor") or "—",
        "principal": info.get("principal") or 0,
        "rate": info.get("rate") or 0,
        "payment": info.get("payment") or 0,
        "term": info.get("term"),
        "notes": info.get("notes") or "",
    }
    aid = _save_pending(context, action)

    preview = _format_pending_action(action)
    warning = ""
    if isinstance(confidence, (int, float)) and confidence < 0.6:
        warning = (
            f"\n\n⚠ <b>Низкая уверенность распознавания ({confidence:.0%}).</b> "
            "Проверьте каждое поле перед сохранением."
        )

    keyboard = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Сохранить", callback_data=f"do:{aid}")],
            [InlineKeyboardButton("❌ Отмена", callback_data=f"cancel:{aid}")],
        ]
    )
    await update.message.reply_text(
        "Из договора получилось:\n\n" + preview + warning + "\n\nСохранить?",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


# ---------------------------------------------------------------------------
# Месячный отчёт в Excel (.xlsx)
# ---------------------------------------------------------------------------


def _build_xlsx_for_chat(cid: int) -> Path:
    """
    Строит .xlsx с отчётом по чату. Возвращает путь к временному файлу.
    Документация openpyxl: https://openpyxl.readthedocs.io/
    """
    if Workbook is None:
        raise RuntimeError("openpyxl не установлен.")

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    # --- Обзор ---
    overview = wb.active
    overview.title = "Обзор"
    inc = total_monthly_income(cid)
    exp = total_monthly_expense(cid)
    debt_pay = total_monthly_debt_payment(cid)
    free = inc - exp - debt_pay
    overview.append(["Показатель", "Значение"])
    style_header(overview)
    overview.append(["Дата отчёта", datetime.now().strftime("%Y-%m-%d %H:%M")])
    overview.append(["Доходы (мес.)", inc])
    overview.append(["Расходы (мес.)", exp])
    overview.append(["Платежи по долгам (мес.)", debt_pay])
    overview.append(["Свободно (мес.)", free])
    overview.column_dimensions["A"].width = 32
    overview.column_dimensions["B"].width = 18

    # --- Доходы ---
    incomes_ws = wb.create_sheet("Доходы")
    incomes_ws.append(["id", "Источник", "Сумма", "Регулярный", "Заметка", "Создано"])
    style_header(incomes_ws)
    with db_conn() as c:
        rows = c.execute(
            "SELECT id, source, amount, is_monthly, notes, created_at "
            "FROM incomes WHERE chat_id=? ORDER BY id", (cid,)
        ).fetchall()
    for r in rows:
        incomes_ws.append(
            [r["id"], r["source"], r["amount"], "да" if r["is_monthly"] else "разовый",
             r["notes"] or "", r["created_at"]]
        )

    # --- Расходы ---
    exp_ws = wb.create_sheet("Расходы")
    exp_ws.append(["id", "Категория", "Сумма", "Регулярный", "Заметка", "Создано"])
    style_header(exp_ws)
    with db_conn() as c:
        rows = c.execute(
            "SELECT id, category, amount, is_monthly, notes, created_at "
            "FROM expenses WHERE chat_id=? ORDER BY id", (cid,)
        ).fetchall()
    for r in rows:
        exp_ws.append(
            [r["id"], r["category"], r["amount"], "да" if r["is_monthly"] else "разовый",
             r["notes"] or "", r["created_at"]]
        )

    # --- Долги ---
    debts_ws = wb.create_sheet("Долги")
    debts_ws.append(["id", "Кредитор", "Остаток", "Ставка %", "Платёж/мес", "Срок (мес.)", "Заметка"])
    style_header(debts_ws)
    for d in fetch_debts(cid):
        debts_ws.append([d.id, d.creditor, d.principal, d.annual_rate_pct,
                         d.monthly_payment, d.term_months or "", d.notes or ""])

    # --- Цели ---
    goals_ws = wb.create_sheet("Цели")
    goals_ws.append(["id", "Название", "Цель", "Накоплено", "% выполнения", "Дедлайн", "Заметка"])
    style_header(goals_ws)
    for g in fetch_goals(cid):
        target = float(g["target"]) if g["target"] else 0
        saved = float(g["saved"]) if g["saved"] else 0
        pct = (saved / target * 100) if target else 0
        goals_ws.append([g["id"], g["name"], target, saved,
                         f"{pct:.0f}%", g["deadline"] or "", g["notes"] or ""])

    for ws in wb.worksheets:
        for col in ws.columns:
            length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(40, max(12, length + 2))

    out = Path(tempfile.gettempdir()) / f"family_finance_report_{cid}_{datetime.now():%Y-%m}.xlsx"
    wb.save(out)
    return out


async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Сформировать и прислать .xlsx-отчёт по требованию."""
    cid = update.effective_chat.id
    await update.message.chat.send_action("upload_document")
    try:
        path = _build_xlsx_for_chat(cid)
    except Exception as e:  # noqa: BLE001
        log.exception("xlsx build error: %s", e)
        await update.message.reply_text(f"Не удалось собрать отчёт: {e}")
        return
    try:
        with open(path, "rb") as f:
            await context.bot.send_document(
                chat_id=cid,
                document=f,
                filename=path.name,
                caption="Текущий отчёт по семейным финансам.",
            )
    finally:
        try:
            path.unlink()
        except OSError:
            pass


def _all_chat_ids() -> list[int]:
    """Возвращает все chat_id, по которым в БД есть данные."""
    with db_conn() as c:
        rows = c.execute(
            """
            SELECT chat_id FROM incomes
            UNION SELECT chat_id FROM expenses
            UNION SELECT chat_id FROM debts
            UNION SELECT chat_id FROM goals
            """
        ).fetchall()
    return [int(r["chat_id"]) for r in rows]


async def monthly_report_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Раз в месяц рассылает .xlsx-отчёт во все чаты с данными."""
    for cid in _all_chat_ids():
        try:
            path = _build_xlsx_for_chat(cid)
        except Exception as e:  # noqa: BLE001
            log.exception("monthly job xlsx error for %s: %s", cid, e)
            continue
        try:
            with open(path, "rb") as f:
                await context.bot.send_document(
                    chat_id=cid,
                    document=f,
                    filename=path.name,
                    caption=f"Месячный отчёт за {datetime.now():%B %Y}.",
                )
        except Exception as e:  # noqa: BLE001
            log.exception("monthly job send error for %s: %s", cid, e)
        finally:
            try:
                path.unlink()
            except OSError:
                pass


# ---------------------------------------------------------------------------
# Запуск
# ---------------------------------------------------------------------------


async def on_unknown(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("Не знаю такой команды. /help — список.")


def main() -> None:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise SystemExit(
            "Не задан TELEGRAM_BOT_TOKEN. Получите токен у @BotFather "
            "(https://core.telegram.org/bots/features#botfather) и экспортируйте его."
        )
    init_db()
    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("add_debt", cmd_add_debt))
    app.add_handler(CommandHandler("list_debts", cmd_list_debts))
    app.add_handler(CommandHandler("del_debt", cmd_del_debt))
    app.add_handler(CommandHandler("add_income", cmd_add_income))
    app.add_handler(CommandHandler("add_expense", cmd_add_expense))
    app.add_handler(CommandHandler("add_goal", cmd_add_goal))
    app.add_handler(CommandHandler("balance", cmd_balance))
    app.add_handler(CommandHandler("analyze", cmd_analyze))
    app.add_handler(CommandHandler("distribute", cmd_distribute))
    app.add_handler(CommandHandler("myid", cmd_myid))
    app.add_handler(CommandHandler("backup", cmd_backup))
    app.add_handler(CommandHandler("report", cmd_report))
    # Подтверждение через inline-кнопки и обработка свободного текста / PDF
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.Document.PDF, on_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    app.add_handler(MessageHandler(filters.COMMAND, on_unknown))

    # Автобэкап раз в BACKUP_INTERVAL_DAYS дней.
    # Документация JobQueue: https://docs.python-telegram-bot.org/en/v21.6/telegram.ext.jobqueue.html
    if app.job_queue is not None and OWNER_TELEGRAM_ID is not None:
        app.job_queue.run_repeating(
            auto_backup_job,
            interval=timedelta(days=BACKUP_INTERVAL_DAYS),
            first=timedelta(minutes=2),  # первый прогон через 2 мин. после старта
            name="auto_backup",
        )
        log.info(
            "Автобэкап включён: раз в %d дн., получатель=%s",
            BACKUP_INTERVAL_DAYS, OWNER_TELEGRAM_ID,
        )
    else:
        log.warning(
            "Автобэкап ВЫКЛЮЧЕН (нет OWNER_TELEGRAM_ID или JobQueue). "
            "Установите python-telegram-bot[job-queue] и задайте OWNER_TELEGRAM_ID."
        )

    # Месячный Excel-отчёт. Выбираем безопасный день (28-е) и фиксированный час.
    # Документация: https://docs.python-telegram-bot.org/en/v21.6/telegram.ext.jobqueue.html
    if app.job_queue is not None and Workbook is not None:
        try:
            app.job_queue.run_monthly(
                monthly_report_job,
                when=time(hour=MONTHLY_REPORT_HOUR, minute=0),
                day=MONTHLY_REPORT_DAY,
                name="monthly_report",
            )
            log.info(
                "Месячный отчёт включён: %d-е число, %02d:00",
                MONTHLY_REPORT_DAY, MONTHLY_REPORT_HOUR,
            )
        except Exception as e:  # noqa: BLE001
            log.warning("Не удалось зарегистрировать месячный отчёт: %s", e)
    else:
        log.warning(
            "Месячный отчёт ВЫКЛЮЧЕН (нет JobQueue или openpyxl). "
            "Команда /report по-прежнему работает по запросу, если openpyxl установлен."
        )

    if llm_available():
        log.info("LLM включён. Модель: %s", CLAUDE_MODEL)
    else:
        log.warning(
            "LLM ВЫКЛЮЧЕН (нет ANTHROPIC_API_KEY или библиотеки anthropic). "
            "Свободный текст и чтение PDF недоступны; команды работают."
        )

    log.info("Бот запущен. Нажмите Ctrl+C для остановки.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
